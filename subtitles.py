import re
import openpyxl

# 读入字幕文件
text = open('10.srt', 'r')

# 空行、行数标号正则表达式
rgx_none_and_num = re.compile(r'\d{1,2}\n')

# 时间正则表达式
rgx_time = re.compile(r'\d\d:\d\d:\d\d,\d\d\d --> \d\d:\d\d:\d\d,\d\d\d\n')

# 处理字幕文件
first_step = text.readlines()

# 新建一个字幕文件
new_file = open('C:\\Users\\18506\\Desktop\\10.srt', 'w')

# 建立4000后的字典
wordlist4001 = {}
excel_content = openpyxl.load_workbook('COCA20000增强版.xlsx')
sheet = excel_content['Sheet1']
for row in range(4002, 17635):
    wordlist4001[sheet.cell(row, 1).value] = sheet.cell(row, 5).value

# 挑选出文字行进行处理
for line in first_step:

    # 如果为空行，行数标号，则不动
    if rgx_none_and_num.search(line):
        new_file.write(line)

    # 如果为时间行则不动
    elif rgx_time.search(line):
        new_file.write(line)

    # 如果为字幕行，则处理
    else:
        line = line.replace('?', ' ')
        line = line.replace(',', ' ')
        line = line.replace('.', ' ')
        words = line.lower().split()
        for word in words:

            # 如果单词不在字典中，则跳过
            if word not in wordlist4001:
                pass
            else:
                new_word = word + ':' + str(wordlist4001[word]) + '\n'
                new_file.write(new_word)
        new_file.write('\n')

# 关闭文件
text.close()
new_file.close()
